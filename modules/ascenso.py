from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from typing import List, Dict, Optional, Any
from datetime import datetime, date
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
import logging
from pathlib import Path
import json
import io
import base64
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time
import math
import os

# Configuración de logging
logger = logging.getLogger(__name__)

# Importar configuración de base de datos
from config import DATABASE_CONFIG as DB_CONFIG

# Configuración de optimización
MAX_WORKERS = 6  # Máximo 6 hilos
BATCH_SIZE = 3000  # Lotes de 3000 registros

router = APIRouter()

# Cargar configuración desde JSON
def cargar_configuracion_ascenso():
    """Carga la configuración de ascenso desde el archivo JSON"""
    try:
        config_path = Path(__file__).parent.parent / "config" / "ascenso_config.json"
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as e:
        logger.error(f"Error cargando configuración de ascenso: {e}")
        # Configuración por defecto en caso de error
        return {
            "niveles_academicos": {
                'BACHILLER': 1,
                'T.S.U': 2,
                'LICENCIATURA': 3,
                'ESPECIALIZACIÓN': 4,
                'MAGISTER': 5,
                'DOCTORADO': 6,
                'POSDOCTORADO': 7,
                'Diplomado en Alta Gerencia': 8
            },
            "rangos": [
                'AGENTE', 'OFICIAL', 'PRIMER OFICIAL', 'OFICIAL JEFE', 
                'INSPECTOR', 'PRIMER INSPECTOR', 'INSPECTOR JEFE', 
                'COMISARIO', 'PRIMER COMISARIO', 'COMISARIO JEFE', 
                'COMISARIO GENERAL', 'COMISARIO MAYOR', 'COMISARIO SUPERIOR'
            ],
            "criterios_ascenso": {}
        }

# Cargar configuración al inicializar el módulo
CONFIG_ASCENSO = cargar_configuracion_ascenso()
NIVELES_ACADEMICOS = CONFIG_ASCENSO["niveles_academicos"]
RANGOS = CONFIG_ASCENSO["rangos"]
CRITERIOS_ASCENSO = CONFIG_ASCENSO["criterios_ascenso"]

# Directorio base para las fotos de funcionarios
FOTOS_BASE_DIR = Path("C:/Users/jozek/Documents/Proyectos/java/jupe/jupe/storage/app/public/fotos_funcionarios/")

# Modelos Pydantic
class FuncionarioAscenso(BaseModel):
    id: int
    cedula: str
    nombre_completo: str
    sexo: Optional[str]
    edad: Optional[int]
    nivel_academico: str
    tiempo_en_rango: float
    tiempo_de_servicio: float
    total_puntos: float
    estado_actual: str
    expedientes: str
    rango_actual: str
    rango_a_aplicar: str
    cumple_todos_requisitos: bool
    cumple_requisitos_menos_academicos: bool
    tiene_expediente_cerrado_reciente: bool
    tiene_expediente_abierto: bool
    condicion_actual_invalida: bool
    observaciones: str
    # Nuevos campos para detalles de expedientes
    detalles_expedientes_abiertos: Optional[str] = ""
    detalles_expedientes_cerrados_recientes: Optional[str] = ""
    expedientes_data: Optional[List[Dict]] = []

class ListadoAscensoRequest(BaseModel):
    fecha_corte: date
    incluir_solo_activos: bool = True
    incluir_solo_uniformados: bool = True

class ListadoAscensoResponse(BaseModel):
    fecha_corte: date
    total_funcionarios_evaluados: int
    listas: Dict[str, List[FuncionarioAscenso]]
    estadisticas: Dict[str, int]
    archivo_excel: Optional[str] = None

def get_db_connection():
    """Crear conexión a la base de datos"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        logger.error(f"Error conectando a la base de datos: {e}")
        raise HTTPException(status_code=500, detail="Error de conexión a la base de datos")

def normalizar_rango(rango: str) -> str:
    """Normaliza el rango actual del funcionario"""
    if not rango or rango.strip() == '':
        return 'NO_REGISTRADO'
    
    rango_limpio = rango.strip().upper()
    
    if rango_limpio not in CRITERIOS_ASCENSO:
        return 'RANGO_INVALIDO'
    
    return rango_limpio

def calcular_edad(fecha_nacimiento: date, fecha_corte: date) -> int:
    """Calcula la edad en años"""
    if not fecha_nacimiento:
        return 0
    
    edad = fecha_corte.year - fecha_nacimiento.year
    if fecha_corte.month < fecha_nacimiento.month or \
       (fecha_corte.month == fecha_nacimiento.month and fecha_corte.day < fecha_nacimiento.day):
        edad -= 1
    
    return edad

def calcular_tiempo_servicio(fecha_ingreso: date, fecha_corte: date, tiempos_adicionales: List[Dict]) -> float:
    """Calcula el tiempo total de servicio en años - CORREGIDO para usar la misma lógica que calcular_antiguedad_policial"""
    if not fecha_ingreso:
        return 0.0
    
    try:
        # Tiempo base desde fecha de ingreso CPNB
        dias_cpnb = (fecha_corte - fecha_ingreso).days
        
        # Tiempo adicional de servicios previos (CORREGIDO: sumar TODOS los tiempos adicionales)
        dias_adicionales = 0
        for tiempo in tiempos_adicionales:
            if tiempo.get('fecha_ingreso') and tiempo.get('fecha_egreso'):
                try:
                    fecha_inicio = datetime.strptime(str(tiempo['fecha_ingreso']), '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(str(tiempo['fecha_egreso']), '%Y-%m-%d').date()
                    # CORREGIDO: Sumar TODOS los períodos de servicio adicional sin restricciones
                    dias_adicionales += (fecha_fin - fecha_inicio).days
                except (ValueError, TypeError):
                    continue
        
        # Total de días combinados
        total_dias = dias_cpnb + dias_adicionales
        
        # Convertir a años (usando la misma lógica que calcular_antiguedad_policial)
        return total_dias / 365.25
        
    except Exception as e:
        logger.error(f"Error calculando tiempo de servicio: {e}")
        return 0.0

def calcular_tiempo_en_rango(fecha_ultimo_ascenso: date, fecha_ingreso: date, fecha_corte: date, historial_ascensos: List[Dict] = None) -> float:
    """
    Calcula el tiempo en el rango actual en años.
    La prioridad de fechas es:
    1. La fecha más reciente del historial de ascensos.
    2. La fecha de último ascenso del funcionario.
    3. La fecha de ingreso del funcionario.
    """
    try:
        fecha_referencia = None
        
        # 1. Prioridad: Historial de Ascensos
        if historial_ascensos:
            fechas_historial = []
            for ascenso in historial_ascensos:
                fecha = ascenso.get('fecha_ascenso')
                if not fecha:
                    continue
                
                if isinstance(fecha, str):
                    try:
                        fecha_dt = datetime.strptime(fecha, '%Y-%m-%d').date()
                        fechas_historial.append(fecha_dt)
                    except (ValueError, TypeError):
                        continue
                elif isinstance(fecha, date):
                    fechas_historial.append(fecha)

            if fechas_historial:
                fecha_referencia = max(fechas_historial)

        # 2. Prioridad: Fecha de último ascenso del funcionario (si no se encontró en el historial)
        if not fecha_referencia or fecha_referencia == date(1900, 1, 1):
            fecha_referencia = fecha_ultimo_ascenso

        # 3. Prioridad: Fecha de ingreso (si ninguna de las anteriores es válida)
        if not fecha_referencia or fecha_referencia == date(1900, 1, 1):
             fecha_referencia = fecha_ingreso

        # Calcular el tiempo si tenemos una fecha de referencia válida
        if fecha_referencia and fecha_referencia != date(1900, 1, 1):
            return (fecha_corte - fecha_referencia).days / 365.25
        
        return 0.0
        
    except Exception as e:
        logger.error(f"Error calculando tiempo en rango: {e}")
        return 0.0

def obtener_nivel_academico_maximo(antecedentes_academicos: List[Dict]) -> str:
    """Obtiene el nivel académico más alto del funcionario"""
    if not antecedentes_academicos:
        return 'NO_REGISTRADO'
    
    nivel_maximo = 0
    grado_maximo = 'NO_REGISTRADO'
    
    for antecedente in antecedentes_academicos:
        grado = str(antecedente.get('grado_instruccion', '')).strip().upper()
        if grado in NIVELES_ACADEMICOS:
            if NIVELES_ACADEMICOS[grado] > nivel_maximo:
                nivel_maximo = NIVELES_ACADEMICOS[grado]
                grado_maximo = grado
    
    return grado_maximo

def calcular_puntos_merito(funcionario_data: Dict) -> float:
    """Calcula los puntos de mérito del funcionario"""
    puntos = 0.0
    
    # Puntos por tiempo de servicio (1 punto por año)
    puntos += funcionario_data.get('tiempo_de_servicio', 0)
    
    # Puntos por nivel académico
    nivel_academico = funcionario_data.get('nivel_academico', 'NO_REGISTRADO')
    if nivel_academico in NIVELES_ACADEMICOS:
        puntos += NIVELES_ACADEMICOS[nivel_academico] * 5  # 5 puntos por nivel
    
    # Puntos por tiempo en rango (0.5 puntos por año)
    puntos += funcionario_data.get('tiempo_en_rango', 0) * 0.5
    
    # Bonificación por no tener expedientes (10 puntos)
    if not funcionario_data.get('tiene_expediente_abierto', False) and \
       not funcionario_data.get('tiene_expediente_cerrado_reciente', False):
        puntos += 10
    
    return round(puntos, 2)

def verificar_expedientes(expedientes: List[Dict], fecha_corte: date, tiempo_requerido_anos: int) -> tuple:
    """Verifica el estado de expedientes del funcionario"""
    tiene_expediente_abierto = False
    tiene_expediente_cerrado_reciente = False
    
    for expediente in expedientes:
        estatus = str(expediente.get('estatus', '')).upper()
        fecha_finalizacion = expediente.get('fecha_finalizacion')
        
        if estatus != 'CERRADO' or not fecha_finalizacion:
            tiene_expediente_abierto = True
        else:
            # Verificar si el expediente cerrado es reciente
            if fecha_finalizacion:
                try:
                    fecha_fin = datetime.strptime(str(fecha_finalizacion), '%Y-%m-%d').date()
                    anos_desde_cierre = (fecha_corte - fecha_fin).days / 365.25
                    
                    if anos_desde_cierre < tiempo_requerido_anos:
                        tiene_expediente_cerrado_reciente = True
                except (ValueError, TypeError):
                    continue
    
    return tiene_expediente_abierto, tiene_expediente_cerrado_reciente

def obtener_funcionarios_para_ascenso(fecha_corte: date) -> List[Dict]:
    """Obtiene todos los funcionarios elegibles para evaluación de ascenso"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Query principal para obtener funcionarios con todas sus relaciones
        # Se agregan fechas máximas para fecha_ingreso y fecha_ultimo_ascenso considerando otras tablas
        query = """
        SELECT 
            f.id,
            f.cedula,
            f.nombre_completo,
            f.sexo,
            f.fecha_nacimiento,
            f.fecha_ingreso,
            f.fecha_ultimo_ascenso,
            f.rango_actual,
            f.status,
            f.condicion_actual,
            f.tipo,
            f.grado_instruccion,
            
            -- Antecedentes académicos
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'grado_instruccion', aa.grado_instruccion,
                        'institucion', aa.institucion,
                        'fecha_graduacion', aa.fecha_graduacion
                    )
                ) FILTER (WHERE aa.id IS NOT NULL), 
                '[]'::json
            ) as antecedentes_academicos,
            
            -- Tiempo de servicio adicional
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'institucion', ts.institucion,
                        'fecha_ingreso', ts.fecha_ingreso,
                        'fecha_egreso', ts.fecha_egreso,
                        'cargo', ts.cargo
                    )
                ) FILTER (WHERE ts.id IS NOT NULL), 
                '[]'::json
            ) as tiempos_servicio,
            
            -- Expedientes
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'nro_expediente', e.nro_expediente,
                        'fecha_inicio', e.fecha_inicio,
                        'fecha_finalizacion', e.fecha_finalizacion,
                        'estatus', e.estatus,
                        'falta', e.falta,
                        'decision', e.decision,
                        'tipo_expediente', e.tipo_expediente,
                        'tipo_sancion_administrativa', e.tipo_sancion_administrativa
                    )
                ) FILTER (WHERE e.id_exp IS NOT NULL), 
                '[]'::json
            ) as expedientes,
            
            -- Historial de ascensos
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'rango_anterior', ha.rango_anterior,
                        'rango_nuevo', ha.rango_nuevo,
                        'fecha_ascenso', ha.fecha_ascenso,
                        'tipo_ascenso', ha.tipo_ascenso
                    )
                ) FILTER (WHERE ha.id IS NOT NULL), 
                '[]'::json
            ) as historial_ascensos
            
        FROM funcionarios f
        LEFT JOIN antecedentes_academicos aa ON f.id = aa.funcionario_id
        LEFT JOIN tiempo_servicio ts ON f.id = ts.funcionario_id
        LEFT JOIN expedientes e ON f.id = e.funcionario_id
        LEFT JOIN historial_ascensos ha ON f.id = ha.funcionario_id
        WHERE 
            f.status = 'ACTIVO'
            AND UPPER(f.tipo) = 'UNIFORMADO'
            AND UPPER(COALESCE(f.condicion_actual, '')) NOT IN ('SOLICITADO', 'DESTITUIDO', 'PRIVADO DE LIBERTAD')
            AND f.rango_actual IS NOT NULL
            AND f.rango_actual != ''
        GROUP BY f.id
        ORDER BY f.cedula
        """
        
        cursor.execute(query)
        funcionarios = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return [dict(funcionario) for funcionario in funcionarios]
        
    except Exception as e:
        logger.error(f"Error obteniendo funcionarios para ascenso: {e}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo funcionarios: {str(e)}")

def procesar_funcionario_para_ascenso(funcionario_data: Dict, fecha_corte: date) -> FuncionarioAscenso:
    """Procesa un funcionario individual para determinar su elegibilidad de ascenso"""
    
    # Datos básicos - Convertir cedula a string si viene como Decimal
    cedula_raw = funcionario_data.get('cedula', '')
    cedula = str(cedula_raw) if cedula_raw else ''
    nombre_completo = funcionario_data.get('nombre_completo', '')
    sexo = funcionario_data.get('sexo', '')
    fecha_nacimiento = funcionario_data.get('fecha_nacimiento')
    fecha_ingreso = funcionario_data.get('fecha_ingreso')
    fecha_ultimo_ascenso = funcionario_data.get('fecha_ultimo_ascenso')
    rango_actual = funcionario_data.get('rango_actual', '')
    condicion_actual = funcionario_data.get('condicion_actual', '')
    
    # Calcular edad
    edad = calcular_edad(fecha_nacimiento, fecha_corte) if fecha_nacimiento else 0
    
    # Normalizar rango
    rango_normalizado = normalizar_rango(rango_actual)
    
    # Obtener criterios de ascenso
    criterios = CRITERIOS_ASCENSO.get(rango_normalizado, {})
    rango_a_aplicar = criterios.get('siguienteRango') or 'N/A'  # Asegurar que nunca sea None
    # Asegurar que los valores nunca sean None para evitar errores de comparación
    tiempo_requerido_rango = criterios.get('tiempoRango') or 0
    antiguedad_requerida = criterios.get('antiguedad') or 0
    nivel_academico_requerido = criterios.get('nivelAcademico', '')
    
    # Procesar antecedentes académicos
    antecedentes = funcionario_data.get('antecedentes_academicos', [])
    if isinstance(antecedentes, str):
        antecedentes = json.loads(antecedentes)
    nivel_academico = obtener_nivel_academico_maximo(antecedentes)
    
    # Procesar tiempos de servicio
    tiempos_servicio = funcionario_data.get('tiempos_servicio', [])
    if isinstance(tiempos_servicio, str):
        tiempos_servicio = json.loads(tiempos_servicio)
    tiempo_de_servicio = calcular_tiempo_servicio(fecha_ingreso, fecha_corte, tiempos_servicio)
    
    # Procesar historial de ascensos
    historial_ascensos = funcionario_data.get('historial_ascensos', [])
    if isinstance(historial_ascensos, str):
        historial_ascensos = json.loads(historial_ascensos)
    
    # Calcular tiempo en rango usando el historial de ascensos
    tiempo_en_rango = calcular_tiempo_en_rango(fecha_ultimo_ascenso, fecha_ingreso, fecha_corte, historial_ascensos)
    
    # Procesar expedientes
    expedientes = funcionario_data.get('expedientes', [])
    if isinstance(expedientes, str):
        expedientes = json.loads(expedientes)
    
    tiene_expediente_abierto, tiene_expediente_cerrado_reciente = verificar_expedientes(
        expedientes, fecha_corte, tiempo_requerido_rango
    )
    
    # Verificar cumplimiento de requisitos
    cumple_tiempo_rango = tiempo_en_rango >= tiempo_requerido_rango
    cumple_antiguedad = tiempo_de_servicio >= antiguedad_requerida
    
    # Corregir lógica académica: puede tener nivel igual o superior al requerido
    cumple_nivel_academico = False
    if nivel_academico_requerido and nivel_academico_requerido in NIVELES_ACADEMICOS:
        if nivel_academico in NIVELES_ACADEMICOS:
            cumple_nivel_academico = NIVELES_ACADEMICOS[nivel_academico] >= NIVELES_ACADEMICOS[nivel_academico_requerido]
    elif not nivel_academico_requerido:  # Si no hay requisito académico específico
        cumple_nivel_academico = True
    
    # Verificar que no tenga expedientes que lo descalifiquen
    sin_expedientes_descalificantes = not tiene_expediente_abierto and not tiene_expediente_cerrado_reciente
    
    cumple_todos_requisitos = (cumple_tiempo_rango and cumple_antiguedad and cumple_nivel_academico and sin_expedientes_descalificantes)
    
    cumple_requisitos_menos_academicos = (cumple_tiempo_rango and cumple_antiguedad and sin_expedientes_descalificantes)
    
    # Verificar condición actual inválida
    condiciones_invalidas = ['SOLICITADO', 'DESTITUIDO', 'PRIVADO DE LIBERTAD']
    condicion_actual_invalida = str(condicion_actual).upper() in condiciones_invalidas
    
    # Calcular datos para el funcionario
    funcionario_procesado = {
        'tiempo_de_servicio': tiempo_de_servicio,
        'tiempo_en_rango': tiempo_en_rango,
        'nivel_academico': nivel_academico,
        'tiene_expediente_abierto': tiene_expediente_abierto,
        'tiene_expediente_cerrado_reciente': tiene_expediente_cerrado_reciente
    }
    
    # Calcular puntos de mérito
    total_puntos = calcular_puntos_merito(funcionario_procesado)
    
    # Procesar detalles de expedientes
    detalles_expedientes_abiertos = []
    detalles_expedientes_cerrados_recientes = []
    
    for expediente in expedientes:
        estatus = str(expediente.get('estatus', '')).upper()
        fecha_finalizacion = expediente.get('fecha_finalizacion')
        nro_expediente = expediente.get('nro_expediente', 'N/A')
        falta = expediente.get('falta', 'N/A')
        decision = expediente.get('decision', 'N/A')
        tipo_expediente = expediente.get('tipo_expediente', 'N/A')
        fecha_inicio = expediente.get('fecha_inicio', 'N/A')
        
        detalle = f"Exp: {nro_expediente}, Tipo: {tipo_expediente}, Falta: {falta}, Decisión: {decision}, Inicio: {fecha_inicio}"
        
        if estatus != 'CERRADO' or not fecha_finalizacion:
            detalles_expedientes_abiertos.append(detalle)
        else:
            # Verificar si el expediente cerrado es reciente
            if fecha_finalizacion:
                try:
                    fecha_fin = datetime.strptime(str(fecha_finalizacion), '%Y-%m-%d').date()
                    anos_desde_cierre = (fecha_corte - fecha_fin).days / 365.25
                    
                    if anos_desde_cierre < tiempo_requerido_rango:
                        detalle_cerrado = f"{detalle}, Cierre: {fecha_finalizacion}, Años desde cierre: {anos_desde_cierre:.1f}"
                        detalles_expedientes_cerrados_recientes.append(detalle_cerrado)
                except (ValueError, TypeError):
                    continue
    
    # Determinar estado de expedientes para mostrar
    if tiene_expediente_abierto:
        estado_expedientes = "EXPEDIENTE ABIERTO"
    elif tiene_expediente_cerrado_reciente:
        estado_expedientes = "EXPEDIENTE CERRADO RECIENTE"
    else:
        estado_expedientes = "SIN EXPEDIENTES"
    
    # Generar observaciones
    observaciones = []
    if not cumple_tiempo_rango:
        observaciones.append(f"Falta tiempo en rango: {tiempo_en_rango:.1f}/{tiempo_requerido_rango} años")
    if not cumple_antiguedad:
        observaciones.append(f"Falta antigüedad: {tiempo_de_servicio:.1f}/{antiguedad_requerida} años")
    if not cumple_nivel_academico:
        observaciones.append(f"Nivel académico insuficiente: {nivel_academico} < {nivel_academico_requerido}")
    if tiene_expediente_abierto:
        observaciones.append("Tiene expediente abierto")
    if tiene_expediente_cerrado_reciente:
        observaciones.append("Expediente cerrado reciente")
    if condicion_actual_invalida:
        observaciones.append(f"Condición actual inválida: {condicion_actual}")
    
    return FuncionarioAscenso(
        id=funcionario_data.get('id'),
        cedula=cedula,
        nombre_completo=nombre_completo,
        sexo=sexo,
        edad=edad,
        nivel_academico=nivel_academico,
        tiempo_en_rango=round(tiempo_en_rango, 2),
        tiempo_de_servicio=round(tiempo_de_servicio, 2),
        total_puntos=total_puntos,
        estado_actual=funcionario_data.get('status', ''),
        expedientes=estado_expedientes,
        rango_actual=rango_actual,
        rango_a_aplicar=rango_a_aplicar,
        cumple_todos_requisitos=cumple_todos_requisitos,
        cumple_requisitos_menos_academicos=cumple_requisitos_menos_academicos,
        tiene_expediente_cerrado_reciente=tiene_expediente_cerrado_reciente,
        tiene_expediente_abierto=tiene_expediente_abierto,
        condicion_actual_invalida=condicion_actual_invalida,
        observaciones="; ".join(observaciones) if observaciones else "Cumple todos los requisitos",
        # Nuevos campos con detalles de expedientes
        detalles_expedientes_abiertos="; ".join(detalles_expedientes_abiertos) if detalles_expedientes_abiertos else "",
        detalles_expedientes_cerrados_recientes="; ".join(detalles_expedientes_cerrados_recientes) if detalles_expedientes_cerrados_recientes else "",
        expedientes_data=expedientes
    )

def calcular_orden_merito_completo(funcionario: FuncionarioAscenso) -> tuple:
    """
    Calcula una tupla de criterios para ordenar funcionarios por mérito completo.
    Retorna tupla con criterios ordenados por prioridad (mayor valor = mejor posición).
    """
    # 1. Tiempo de servicio (años) - Criterio principal
    tiempo_servicio = funcionario.tiempo_de_servicio
    
    # 2. Nivel académico (valor numérico) - Criterio secundario
    nivel_academico_valor = NIVELES_ACADEMICOS.get(funcionario.nivel_academico, 0)
    
    # 3. Tiempo en rango (años) - Criterio terciario
    tiempo_rango = funcionario.tiempo_en_rango
    
    # 4. Puntos de mérito totales - Criterio cuaternario
    puntos_merito = funcionario.total_puntos
    
    # 5. Edad (como desempate final, menor edad = mejor)
    edad_desempate = -(funcionario.edad or 0)  # Negativo para que menor edad tenga mayor valor
    
    return (tiempo_servicio, nivel_academico_valor, tiempo_rango, puntos_merito, edad_desempate)

def organizar_listas_ascenso(funcionarios_procesados: List[FuncionarioAscenso]) -> Dict[str, List[FuncionarioAscenso]]:
    """Organiza los funcionarios en las diferentes listas según los criterios de mérito"""
    
    listas = {
        "cumple_todos_requisitos": [],
        "falta_nivel_academico": [],
        "falta_tiempo_rango": [],
        "falta_tiempo_servicio": [],
        "expediente_cerrado_reciente": [],
        "expediente_abierto": [],
        "condicion_actual_invalida": []
    }
    
    for funcionario in funcionarios_procesados:
        # Verificar condición actual inválida primero (descalifica completamente)
        if funcionario.condicion_actual_invalida:
            listas["condicion_actual_invalida"].append(funcionario)
            continue
            
        # Verificar expedientes que descalifican
        if funcionario.tiene_expediente_abierto:
            listas["expediente_abierto"].append(funcionario)
            continue
            
        if funcionario.tiene_expediente_cerrado_reciente:
            listas["expediente_cerrado_reciente"].append(funcionario)
            continue
        
        # Ahora clasificar por requisitos faltantes (orden de mérito)
        
        # 1. Primera prioridad: Cumple todos los requisitos
        if funcionario.cumple_todos_requisitos:
            listas["cumple_todos_requisitos"].append(funcionario)
        
        # 2. Segunda prioridad: Solo falta nivel académico
        elif funcionario.cumple_requisitos_menos_academicos:
            listas["falta_nivel_academico"].append(funcionario)
        
        # 3-4. Determinar si falta tiempo en rango o tiempo de servicio
        else:
            # Obtener criterios para determinar qué falta específicamente
            rango_normalizado = normalizar_rango(funcionario.rango_actual)
            criterios = CRITERIOS_ASCENSO.get(rango_normalizado, {})
            tiempo_requerido_rango = criterios.get('tiempoRango') or 0
            antiguedad_requerida = criterios.get('antiguedad') or 0
            
            cumple_tiempo_rango = funcionario.tiempo_en_rango >= tiempo_requerido_rango
            cumple_antiguedad = funcionario.tiempo_de_servicio >= antiguedad_requerida
            
            # 3. Tercera prioridad: Solo falta tiempo en rango
            if cumple_antiguedad and not cumple_tiempo_rango:
                listas["falta_tiempo_rango"].append(funcionario)
            
            # 4. Cuarta prioridad: Falta tiempo de servicio (con o sin tiempo en rango)
            elif not cumple_antiguedad:
                listas["falta_tiempo_servicio"].append(funcionario)
            
            # Caso por defecto (no debería llegar aquí, pero por seguridad)
            else:
                listas["falta_tiempo_servicio"].append(funcionario)
    
    # Ordenar cada lista por orden de mérito completo (múltiples criterios)
    logger.info("Aplicando ordenamiento por mérito completo a cada categoría...")
    
    for lista_nombre in listas:
        if listas[lista_nombre]:  # Solo ordenar si hay funcionarios
            # Ordenar por criterios múltiples: tiempo servicio, nivel académico, tiempo rango, puntos, edad
            listas[lista_nombre].sort(key=calcular_orden_merito_completo, reverse=True)
            
            # Log para verificar el orden
            logger.info(f"Categoría '{lista_nombre}': {len(listas[lista_nombre])} funcionarios ordenados por mérito")
            if len(listas[lista_nombre]) > 0:
                primer_funcionario = listas[lista_nombre][0]
                logger.debug(f"  Primer funcionario: {primer_funcionario.nombre_completo} - "
                           f"Servicio: {primer_funcionario.tiempo_de_servicio:.1f}años, "
                           f"Académico: {primer_funcionario.nivel_academico}, "
                           f"Rango: {primer_funcionario.tiempo_en_rango:.1f}años, "
                           f"Puntos: {primer_funcionario.total_puntos}")
    
    return listas

def obtener_datos_funcionario_completos(funcionario_id: int) -> Dict:
    """Obtiene todos los datos completos de un funcionario para el listado de ascenso"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Query completa para obtener todos los datos necesarios
        query = """
        SELECT 
            f.id,
            f.cedula,
            f.nombre_completo,
            f.sexo,
            f.fecha_nacimiento,
            f.fecha_ingreso,
            f.fecha_ultimo_ascenso,
            f.rango_actual,
            f.status,
            f.condicion_actual,
            f.tipo,
            f.grado_instruccion,
            f.nro_local,
            f.nro_celular,
            f.nro_otro,
            f.cohorte,
            f.imagen,
            f.dependencia as dependencia_texto,
            
            -- Dependencia con prioridad (pivot sobre funcionario)
            COALESCE(
                (SELECT d.nombre 
                 FROM funcionario_dependencia fd 
                 JOIN dependencias d ON fd.id_dependencia = d.id_dependencia 
                 WHERE fd.id_funcionario = f.id 
                 ORDER BY fd.created_at DESC 
                 LIMIT 1),
                f.dependencia,
                'No posee'
            ) as dependencia_final,
            
            -- Rango actual con prioridad (historial sobre funcionario)
            COALESCE(
                (SELECT ha.rango_nuevo 
                 FROM historial_ascensos ha 
                 WHERE ha.funcionario_id = f.id 
                 ORDER BY ha.fecha_ascenso DESC 
                 LIMIT 1),
                f.rango_actual
            ) as rango_final,
            
            -- Fecha último ascenso con prioridad (historial sobre funcionario)
            COALESCE(
                (SELECT ha.fecha_ascenso 
                 FROM historial_ascensos ha 
                 WHERE ha.funcionario_id = f.id 
                 ORDER BY ha.fecha_ascenso DESC 
                 LIMIT 1),
                f.fecha_ultimo_ascenso
            ) as fecha_ultimo_ascenso_final,
            
            -- Nivel académico con prioridad (antecedentes sobre funcionario)
            COALESCE(
                (SELECT aa.grado_instruccion 
                 FROM antecedentes_academicos aa 
                 WHERE aa.funcionario_id = f.id 
                 ORDER BY aa.fecha_graduacion DESC 
                 LIMIT 1),
                f.grado_instruccion
            ) as nivel_academico_final,
            
            -- Observaciones del último ascenso
            (SELECT ha.observaciones 
             FROM historial_ascensos ha 
             WHERE ha.funcionario_id = f.id 
             ORDER BY ha.fecha_ascenso DESC 
             LIMIT 1) as observaciones_ascenso,
            
            -- Tiempo de servicio adicional
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'institucion', ts.institucion,
                        'fecha_ingreso', ts.fecha_ingreso,
                        'fecha_egreso', ts.fecha_egreso,
                        'cargo', ts.cargo
                    )
                ) FILTER (WHERE ts.id IS NOT NULL), 
                '[]'::json
            ) as tiempos_servicio
            
        FROM funcionarios f
        LEFT JOIN tiempo_servicio ts ON f.id = ts.funcionario_id
        WHERE f.id = %s
        GROUP BY f.id
        """
        
        cursor.execute(query, (funcionario_id,))
        funcionario_data = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        return dict(funcionario_data) if funcionario_data else {}
        
    except Exception as e:
        # Solo loggear errores críticos, no errores de datos faltantes
        if "no existe la columna" not in str(e).lower() and "does not exist" not in str(e).lower():
            logger.error(f"Error obteniendo datos completos del funcionario {funcionario_id}: {e}")
        return {}

def calcular_antiguedad_policial(fecha_ingreso: date, tiempos_servicio: List[Dict], fecha_corte: date) -> str:
    """Calcula la antigüedad total en la función policial como en show.blade.php"""
    if not fecha_ingreso:
        return "0 años, 0 meses, 0 días"
    
    try:
        # Tiempo base desde fecha de ingreso CPNB
        dias_cpnb = (fecha_corte - fecha_ingreso).days
        
        # Tiempo adicional de servicios previos
        dias_adicionales = 0
        for tiempo in tiempos_servicio:
            if tiempo.get('fecha_ingreso') and tiempo.get('fecha_egreso'):
                try:
                    fecha_inicio = datetime.strptime(str(tiempo['fecha_ingreso']), '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(str(tiempo['fecha_egreso']), '%Y-%m-%d').date()
                    dias_adicionales += (fecha_fin - fecha_inicio).days
                except (ValueError, TypeError):
                    continue
        
        # Total de días combinados
        total_dias = dias_cpnb + dias_adicionales
        
        # Convertir a años, meses y días
        total_anos = total_dias // 365
        dias_restantes = total_dias % 365
        total_meses = dias_restantes // 30
        dias_finales = dias_restantes % 30
        
        return f"{total_anos} años, {total_meses} meses, {dias_finales} días"
        
    except Exception as e:
        logger.error(f"Error calculando antigüedad policial: {e}")
        return "Error en cálculo"

def procesar_lote_funcionarios_excel(lote_funcionarios: List[FuncionarioAscenso], fecha_corte: date, lote_id: int) -> List[Dict]:
    """Procesa un lote de funcionarios para Excel en un hilo separado"""
    start_time = time.time()
    filas_procesadas = []
    
    logger.info(f"[EXCEL_LOTE_{lote_id}] Iniciando procesamiento de {len(lote_funcionarios)} funcionarios")
    
    # Obtener conexión a la base de datos para este hilo
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Preparar query optimizada para obtener datos de múltiples funcionarios
        funcionario_ids = [f.id for f in lote_funcionarios]
        placeholders = ','.join(['%s'] * len(funcionario_ids))
        
        query = f"""
        SELECT 
            f.id,
            f.cedula,
            f.nombre_completo,
            f.sexo,
            f.fecha_nacimiento,
            f.fecha_ingreso,
            f.fecha_ultimo_ascenso,
            f.rango_actual,
            f.status,
            f.condicion_actual,
            f.tipo,
            f.grado_instruccion,
            f.nro_local,
            f.nro_celular,
            f.nro_otro,
            f.cohorte,
            f.imagen,
            f.dependencia as dependencia_texto,
            
            -- Dependencia con prioridad
            COALESCE(
                (SELECT d.nombre 
                 FROM funcionario_dependencia fd 
                 JOIN dependencias d ON fd.id_dependencia = d.id_dependencia 
                 WHERE fd.id_funcionario = f.id 
                 ORDER BY fd.created_at DESC 
                 LIMIT 1),
                f.dependencia,
                'No posee'
            ) as dependencia_final,
            
            -- Rango actual con prioridad
            COALESCE(
                (SELECT ha.rango_nuevo 
                 FROM historial_ascensos ha 
                 WHERE ha.funcionario_id = f.id 
                 ORDER BY ha.fecha_ascenso DESC 
                 LIMIT 1),
                f.rango_actual
            ) as rango_final,
            
            -- Fecha último ascenso con prioridad
            COALESCE(
                (SELECT ha.fecha_ascenso 
                 FROM historial_ascensos ha 
                 WHERE ha.funcionario_id = f.id 
                 ORDER BY ha.fecha_ascenso DESC 
                 LIMIT 1),
                f.fecha_ultimo_ascenso
            ) as fecha_ultimo_ascenso_final,
            
            -- Nivel académico con prioridad
            COALESCE(
                (SELECT aa.grado_instruccion 
                 FROM antecedentes_academicos aa 
                 WHERE aa.funcionario_id = f.id 
                 ORDER BY aa.fecha_graduacion DESC 
                 LIMIT 1),
                f.grado_instruccion
            ) as nivel_academico_final,
            
            -- Observaciones del último ascenso
            (SELECT ha.observaciones 
             FROM historial_ascensos ha 
             WHERE ha.funcionario_id = f.id 
             ORDER BY ha.fecha_ascenso DESC 
             LIMIT 1) as observaciones_ascenso,
            
            -- Tiempo de servicio adicional
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'institucion', ts.institucion,
                        'fecha_ingreso', ts.fecha_ingreso,
                        'fecha_egreso', ts.fecha_egreso,
                        'cargo', ts.cargo
                    )
                ) FILTER (WHERE ts.id IS NOT NULL), 
                '[]'::json
            ) as tiempos_servicio
            
        FROM funcionarios f
        LEFT JOIN tiempo_servicio ts ON f.id = ts.funcionario_id
        WHERE f.id IN ({placeholders})
        GROUP BY f.id
        ORDER BY f.id
        """
        
        cursor.execute(query, funcionario_ids)
        datos_funcionarios = {row['id']: dict(row) for row in cursor.fetchall()}
        
        cursor.close()
        conn.close()
        
        # Procesar cada funcionario del lote
        for funcionario in lote_funcionarios:
            try:
                datos_completos = datos_funcionarios.get(funcionario.id, {})
                
                if not datos_completos:
                    # Usar datos básicos del funcionario procesado
                    datos_completos = {
                        'nombre_completo': funcionario.nombre_completo,
                        'cedula': funcionario.cedula,
                        'dependencia_final': 'No posee',
                        'nivel_academico_final': funcionario.nivel_academico,
                        'cohorte': 'Sin registrar',
                        'fecha_ingreso': None,
                        'fecha_ultimo_ascenso_final': None,
                        'observaciones_ascenso': funcionario.observaciones,
                        'nro_celular': None,
                        'nro_local': None,
                        'nro_otro': None,
                        'imagen': None,
                        'rango_final': funcionario.rango_actual,
                        'tiempos_servicio': []
                    }
                
                # Procesar tiempos de servicio
                tiempos_servicio = datos_completos.get('tiempos_servicio', [])
                if isinstance(tiempos_servicio, str):
                    try:
                        tiempos_servicio = json.loads(tiempos_servicio)
                    except:
                        tiempos_servicio = []
                
                # Calcular antigüedad policial
                antiguedad_policial = calcular_antiguedad_policial(
                    datos_completos.get('fecha_ingreso'),
                    tiempos_servicio,
                    fecha_corte
                )
                
                # Preparar teléfono
                telefono = datos_completos.get('nro_celular') or datos_completos.get('nro_local') or datos_completos.get('nro_otro') or 'Sin registrar'
                
                # Preparar fecha último ascenso
                fecha_ultimo_ascenso = datos_completos.get('fecha_ultimo_ascenso_final')
                if fecha_ultimo_ascenso:
                    try:
                        fecha_ultimo_ascenso_str = fecha_ultimo_ascenso.strftime('%d-%m-%Y') if isinstance(fecha_ultimo_ascenso, date) else str(fecha_ultimo_ascenso)
                    except:
                        fecha_ultimo_ascenso_str = "NO POSEE"
                else:
                    fecha_ultimo_ascenso_str = "NO POSEE"
                
                # Preparar fecha ingreso
                fecha_ingreso = datos_completos.get('fecha_ingreso')
                try:
                    fecha_ingreso_str = fecha_ingreso.strftime('%d-%m-%Y') if fecha_ingreso else 'Sin registrar'
                except:
                    fecha_ingreso_str = 'Sin registrar'
                
                # Crear datos del funcionario para la columna B (uno debajo del otro)
                datos_funcionario = [
                    datos_completos.get('nombre_completo', ''),
                    datos_completos.get('cedula', ''),
                    str(funcionario.edad or 0),
                    datos_completos.get('rango_final', '') or funcionario.rango_actual,
                    telefono,
                    datos_completos.get('dependencia_final', '') or 'Sin registrar',
                    datos_completos.get('nivel_academico_final', '') or funcionario.nivel_academico,
                    datos_completos.get('cohorte', '') or 'Sin registrar',
                    fecha_ingreso_str,
                    fecha_ultimo_ascenso_str,
                    antiguedad_policial
                ]
                
                # Crear fila de datos con el nuevo formato
                fila_data = {
                    'funcionario_id': funcionario.id,
                    'datos_funcionario': datos_funcionario,  # Lista de datos para la columna B
                    'rango_a_aplicar': funcionario.rango_a_aplicar,  # Rango al que aplica para ascender
                    'observaciones': "",  # En blanco para que el evaluador escriba
                    'requisitos_no_cumplidos': funcionario.observaciones,  # Lo que antes iba en observaciones
                    'apto': "",  # Dropdown
                    'no_apto': "",  # Dropdown
                    'tiene_imagen': bool(datos_completos.get('imagen')),
                    'imagen_nombre': datos_completos.get('imagen', '')
                }
                
                filas_procesadas.append(fila_data)
                
            except Exception as row_error:
                logger.error(f"[EXCEL_LOTE_{lote_id}] Error procesando funcionario {funcionario.cedula}: {row_error}")
                continue
        
        processing_time = time.time() - start_time
        logger.info(f"[EXCEL_LOTE_{lote_id}] Completado en {processing_time:.2f}s - {len(filas_procesadas)} filas procesadas")
        
        return filas_procesadas
        
    except Exception as e:
        logger.error(f"[EXCEL_LOTE_{lote_id}] Error en procesamiento: {e}")
        return []

def generar_excel_ascensos(listas: Dict[str, List[FuncionarioAscenso]], fecha_corte: date) -> str:
    """Genera un archivo Excel con las listas de ascenso usando procesamiento multihilo optimizado"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from PIL import Image as PILImage
        import tempfile
        import os
        
        logger.info("Iniciando generación de archivo Excel con procesamiento multihilo...")
        
        # Contar total de funcionarios a procesar
        total_funcionarios = sum(len(funcionarios) for funcionarios in listas.values())
        logger.info(f"Total de funcionarios a incluir en Excel: {total_funcionarios}")
        
        # Crear lista ordenada de funcionarios respetando el orden de mérito
        todos_funcionarios = []
        
        # Orden de prioridad según mérito policial
        orden_listas = [
            "cumple_todos_requisitos",
            "falta_nivel_academico", 
            "falta_tiempo_rango",
            "falta_tiempo_servicio",
            "expediente_cerrado_reciente",
            "expediente_abierto",
            "condicion_actual_invalida"
        ]
        
        for categoria in orden_listas:
            if categoria in listas:
                funcionarios_categoria = listas[categoria]
                logger.info(f"Agregando {len(funcionarios_categoria)} funcionarios de categoría: {categoria}")
                todos_funcionarios.extend(funcionarios_categoria)
        
        logger.info(f"Procesando {len(todos_funcionarios)} funcionarios en lotes de {BATCH_SIZE}")
        
        # Dividir funcionarios en lotes manteniendo el orden
        lotes = []
        for i in range(0, len(todos_funcionarios), BATCH_SIZE):
            lote = todos_funcionarios[i:i + BATCH_SIZE]
            lotes.append(lote)
        
        # Procesar lotes en paralelo PERO manteniendo el orden de mérito
        todas_filas = []
        start_time = time.time()
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Enviar lotes a los hilos con índice para mantener orden
            futures = []
            for lote_id, lote in enumerate(lotes):
                future = executor.submit(procesar_lote_funcionarios_excel, lote, fecha_corte, lote_id)
                futures.append((lote_id, future))
            
            # Recopilar resultados EN ORDEN para mantener el mérito policial
            resultados_ordenados = [None] * len(lotes)  # Array para mantener orden
            
            for lote_id, future in futures:
                try:
                    filas_lote = future.result()
                    resultados_ordenados[lote_id] = filas_lote  # Guardar en posición correcta
                    logger.info(f"[EXCEL_MULTIHILO] Lote {lote_id} completado - {len(filas_lote)} filas")
                except Exception as e:
                    logger.error(f"[EXCEL_MULTIHILO] Error en lote {lote_id}: {e}")
                    resultados_ordenados[lote_id] = []  # Lista vacía en caso de error
            
            # Concatenar resultados en el orden correcto de mérito
            for resultado_lote in resultados_ordenados:
                if resultado_lote:  # Solo agregar si no es None o vacío
                    todas_filas.extend(resultado_lote)
        
        processing_time = time.time() - start_time
        logger.info(f"[EXCEL_MULTIHILO] Procesamiento completado en {processing_time:.2f}s")
        logger.info(f"[EXCEL_MULTIHILO] Total filas procesadas: {len(todas_filas)}")
        
        # Crear workbook
        logger.info("Creando archivo Excel...")
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado de Ascensos"
        
        # Crear hoja de instrucciones separada
        ws_instrucciones = wb.create_sheet("Instrucciones")
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Configurar hoja de instrucciones
        logger.info("Creando hoja de instrucciones...")
        ws_instrucciones.column_dimensions['A'].width = 30
        ws_instrucciones.column_dimensions['B'].width = 80
        
        # Título de instrucciones
        title_cell = ws_instrucciones.cell(row=1, column=1, value="INSTRUCCIONES DE USO - LISTADO DE ASCENSOS")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_instrucciones.merge_cells('A1:B1')
        ws_instrucciones.row_dimensions[1].height = 30
        
        # Instrucciones detalladas
        instrucciones_detalladas = [
            ("", ""),  # Fila vacía
            ("COLUMNA", "DESCRIPCIÓN"),
            ("Foto", "Muestra la fotografía del funcionario si está disponible"),
            ("Datos del Funcionario", "Contiene toda la información personal y profesional del funcionario:\n• Nombre Completo\n• Cédula\n• Edad\n• Rango Actual\n• Teléfono\n• Dependencia\n• Nivel Académico\n• Cohorte\n• Fecha Ingreso\n• Fecha Último Ascenso\n• Antigüedad en Función Policial"),
            ("Observaciones", "CAMPO PARA LLENAR: Escriba aquí sus observaciones sobre la evaluación del funcionario"),
            ("Requisitos No Cumplidos", "Muestra automáticamente los requisitos que no cumple el funcionario según los criterios de ascenso"),
            ("Apto", "CAMPO PARA LLENAR: Marque 'Sí' si el funcionario ES APTO para ascenso"),
            ("No Apto", "CAMPO PARA LLENAR: Marque 'Sí' si el funcionario NO ES APTO para ascenso"),
            ("", ""),  # Fila vacía
            ("IMPORTANTE", "NOTAS IMPORTANTES"),
            ("Orden de Mérito", "Los funcionarios están ordenados por orden de mérito policial:\n1. Cumple todos los requisitos\n2. Solo falta nivel académico\n3. Solo falta tiempo en rango\n4. Solo falta tiempo de servicio\n5. Expediente cerrado reciente\n6. Expediente abierto\n7. Condición actual inválida"),
            ("Evaluación", "Solo marque UNA opción: 'Apto' O 'No Apto', nunca ambas"),
            ("Navegación", "Use los filtros de Excel para navegar por los datos sin alterar el orden de mérito"),
            ("Guardado", "Guarde el archivo frecuentemente para no perder su trabajo")
        ]
        
        for row_idx, (columna, descripcion) in enumerate(instrucciones_detalladas, 2):
            if columna == "COLUMNA" or columna == "IMPORTANTE":
                # Encabezados de sección
                cell_a = ws_instrucciones.cell(row=row_idx, column=1, value=columna)
                cell_b = ws_instrucciones.cell(row=row_idx, column=2, value=descripcion)
                cell_a.font = Font(bold=True, color="FFFFFF")
                cell_b.font = Font(bold=True, color="FFFFFF")
                cell_a.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell_b.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            elif columna == "":
                # Filas vacías
                continue
            else:
                # Contenido normal
                cell_a = ws_instrucciones.cell(row=row_idx, column=1, value=columna)
                cell_b = ws_instrucciones.cell(row=row_idx, column=2, value=descripcion)
                cell_a.font = Font(bold=True)
                cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Aplicar bordes
            cell_a.border = border
            cell_b.border = border
            
            # Ajustar altura de fila para texto largo
            if "\n" in descripcion:
                ws_instrucciones.row_dimensions[row_idx].height = 80
            else:
                ws_instrucciones.row_dimensions[row_idx].height = 25
        
        # Configurar hoja principal
        logger.info("Configurando hoja principal...")
        
        # Encabezados de columnas con el nuevo formato
        headers = [
            "Foto", 
            "Datos del Funcionario", 
            "Observaciones", 
            "Requisitos No Cumplidos", 
            "Apto", 
            "No Apto"
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        column_widths = [15, 50, 40, 40, 10, 10]  # Foto, Datos, Observaciones, Requisitos, Apto, No Apto
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Configurar altura de filas para las fotos
        ws.row_dimensions[1].height = 30  # Header
        
        # Escribir datos de las filas
        logger.info("Escribiendo datos al Excel...")
        current_row = 2  # Empezar en la fila 2 directamente después del encabezado
        
        # Etiquetas para los datos del funcionario
        etiquetas_datos = [
            "Nombre Completo:",
            "Cédula:",
            "Edad:",
            "Rango Actual:",
            "Teléfono:",
            "Dependencia:",
            "Nivel Académico:",
            "Cohorte:",
            "Fecha Ingreso:",
            "Fecha Último Ascenso:",
            "Antigüedad en Función Policial:"
        ]
        
        for fila_data in todas_filas:
            try:
                # Configurar altura de fila (más alta para acomodar todos los datos)
                if fila_data.get('tiene_imagen'):
                    ws.row_dimensions[current_row].height = 200  # Altura mayor para fotos y datos
                else:
                    ws.row_dimensions[current_row].height = 180  # Altura para datos sin foto
                
                # Columna A: Foto
                if fila_data.get('tiene_imagen'):
                    try:
                        imagen_nombre = fila_data.get('imagen_nombre', '')
                        if imagen_nombre:
                            # Construir ruta de la imagen según el patrón encontrado
                            # El campo imagen contiene algo como: "98366_12115547/fotos/687d8a72b5615.png"
                            # Y debe buscarse en: storage/app/public/fotos_funcionarios/fotos/98366_12115547/fotos/687d8a72b5615.png
                            
                            # Rutas alternativas a intentar basadas en la estructura real
                            # El campo imagen contiene: "98366_12115547/fotos/687d8a72b5615.png"
                            # Debe buscarse en: storage/app/public/fotos_funcionarios/{imagen_nombre}
                            rutas_alternativas = [
                                f"storage/app/public/fotos_funcionarios/{imagen_nombre}",      # Ruta principal según especificación
                                f"storage/app/public/fotos_funcionarios/fotos/{imagen_nombre}",  # Ruta alternativa 1
                                f"public/storage/fotos_funcionarios/{imagen_nombre}",         # Ruta alternativa 2
                                f"public/storage/fotos_funcionarios/fotos/{imagen_nombre}",    # Ruta alternativa 3
                                str(FOTOS_BASE_DIR / imagen_nombre)                           # Ruta con Path
                            ]
                            
                            imagen_encontrada = False
                            for ruta in rutas_alternativas:
                                if os.path.exists(ruta):
                                    imagen_path = ruta
                                    imagen_encontrada = True
                                    logger.debug(f"Imagen encontrada en: {ruta}")
                                    break
                            
                            if imagen_encontrada:
                                # Redimensionar imagen
                                with PILImage.open(imagen_path) as img:
                                    # Redimensionar manteniendo proporción
                                    img.thumbnail((120, 160), PILImage.Resampling.LANCZOS)
                                    
                                    # Crear archivo temporal
                                    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                                        img.save(temp_file.name, 'PNG')
                                        
                                        # Agregar imagen a Excel
                                        excel_img = Image(temp_file.name)
                                        excel_img.width = 100
                                        excel_img.height = 140
                                        
                                        # Posicionar imagen en la celda
                                        cell_ref = f"A{current_row}"
                                        ws.add_image(excel_img, cell_ref)
                                        
                                        # Limpiar archivo temporal
                                        os.unlink(temp_file.name)
                            else:
                                # Si no se encuentra la imagen, escribir texto con información de debug
                                logger.warning(f"Imagen no encontrada: {imagen_nombre}. Rutas intentadas: {rutas_alternativas}")
                                cell = ws.cell(row=current_row, column=1, value="Foto no encontrada")
                                cell.border = border
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception as img_error:
                        logger.warning(f"Error procesando imagen para funcionario ID {fila_data['funcionario_id']}: {img_error}")
                        # Escribir texto si hay error con la imagen
                        cell = ws.cell(row=current_row, column=1, value="Error cargando foto")
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Sin imagen
                    cell = ws.cell(row=current_row, column=1, value="Sin foto")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Columna B: Datos del Funcionario (uno debajo del otro) + Rango a aplicar
                datos_texto = []
                for i, (etiqueta, valor) in enumerate(zip(etiquetas_datos, fila_data['datos_funcionario'])):
                    datos_texto.append(f"{etiqueta} {valor}")
                
                # Agregar el rango a aplicar al final como parte del texto normal
                rango_a_aplicar = fila_data.get('rango_a_aplicar', 'N/A')
                datos_texto.append(f"Rango a Aplicar: {rango_a_aplicar}")
                
                datos_completos = "\n".join(datos_texto)
                cell = ws.cell(row=current_row, column=2, value=datos_completos)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Columna C: Observaciones (en blanco para que el evaluador escriba)
                cell = ws.cell(row=current_row, column=3, value=fila_data['observaciones'])
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Columna D: Requisitos No Cumplidos
                cell = ws.cell(row=current_row, column=4, value=fila_data['requisitos_no_cumplidos'])
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Columna E: Apto
                cell = ws.cell(row=current_row, column=5, value=fila_data['apto'])
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Columna F: No Apto
                cell = ws.cell(row=current_row, column=6, value=fila_data['no_apto'])
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
                
            except Exception as row_error:
                logger.error(f"Error escribiendo fila: {row_error}")
                continue
        
        logger.info(f"Completado escritura de {current_row - 2} filas de datos")
        
        # Agregar validación de datos para las columnas Apto/No Apto
        if current_row > 2:  # Solo si hay datos
            try:
                logger.info("Agregando validaciones de datos...")
                
                apto_validation = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
                apto_validation.error = 'Seleccione Sí o No para indicar si es apto'
                apto_validation.errorTitle = 'Entrada inválida'
                
                # Aplicar validación a la columna E (Apto) - empezar desde fila 2
                ws.add_data_validation(apto_validation)
                apto_validation.add(f"E2:E{current_row-1}")
                
                no_apto_validation = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
                no_apto_validation.error = 'Seleccione Sí o No para indicar si NO es apto'
                no_apto_validation.errorTitle = 'Entrada inválida'
                
                ws.add_data_validation(no_apto_validation)
                no_apto_validation.add(f"F2:F{current_row-1}")
                
                # Congelar primera fila (solo encabezado)
                ws.freeze_panes = "A2"
                
                # Agregar autofiltro desde la fila 1
                ws.auto_filter.ref = f"A1:F{current_row-1}"
                
                logger.info("Validaciones agregadas exitosamente")
            except Exception as validation_error:
                logger.warning(f"Error agregando validaciones: {validation_error}")
        
        logger.info("Guardando archivo Excel en buffer...")
        
        # Guardar en buffer
        wb.save(output)
        
        # Obtener el contenido del buffer
        output.seek(0)
        excel_content = output.getvalue()
        
        logger.info(f"Excel generado exitosamente. Tamaño: {len(excel_content)} bytes")
        
        # Codificar en base64 para envío
        excel_base64 = base64.b64encode(excel_content).decode('utf-8')
        
        logger.info("Excel codificado en base64 exitosamente")
        
        return excel_base64
        
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando archivo Excel: {str(e)}")

# Endpoints

@router.get("/ascenso/criterios")
async def get_ascenso_criterios():
    """
    Endpoint para obtener los criterios de ascenso, rangos y niveles académicos.
    """
    return {
        "nivelesAcademicos": NIVELES_ACADEMICOS,
        "rangos": RANGOS,
        "criteriosAscenso": CRITERIOS_ASCENSO
    }

def procesar_lote_funcionarios(lote_funcionarios: List[Dict], fecha_corte: date, lote_id: int) -> List[FuncionarioAscenso]:
    """Procesa un lote de funcionarios en un hilo separado"""
    start_time = time.time()
    funcionarios_procesados = []
    
    logger.info(f"[LOTE_{lote_id}] Iniciando procesamiento de {len(lote_funcionarios)} funcionarios")
    
    for funcionario_data in lote_funcionarios:
        try:
            funcionario_procesado = procesar_funcionario_para_ascenso(funcionario_data, fecha_corte)
            funcionarios_procesados.append(funcionario_procesado)
        except Exception as e:
            logger.error(f"[LOTE_{lote_id}] Error procesando funcionario {funcionario_data.get('cedula', 'N/A')}: {e}")
            continue
    
    processing_time = time.time() - start_time
    logger.info(f"[LOTE_{lote_id}] Completado en {processing_time:.2f}s - {len(funcionarios_procesados)} funcionarios procesados")
    
    return funcionarios_procesados

def procesar_funcionarios_multihilo(funcionarios_data: List[Dict], fecha_corte: date) -> List[FuncionarioAscenso]:
    """Procesa funcionarios usando múltiples hilos y lotes"""
    total_funcionarios = len(funcionarios_data)
    
    if total_funcionarios == 0:
        return []
    
    # Calcular número de lotes
    num_lotes = math.ceil(total_funcionarios / BATCH_SIZE)
    
    logger.info(f"[MULTIHILO] Procesando {total_funcionarios} funcionarios en {num_lotes} lotes de {BATCH_SIZE}")
    logger.info(f"[MULTIHILO] Usando máximo {MAX_WORKERS} hilos")
    
    # Dividir funcionarios en lotes
    lotes = []
    for i in range(0, total_funcionarios, BATCH_SIZE):
        lote = funcionarios_data[i:i + BATCH_SIZE]
        lotes.append(lote)
    
    # Procesar lotes en paralelo
    funcionarios_procesados = []
    start_time = time.time()
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Enviar lotes a los hilos
        future_to_lote = {}
        for lote_id, lote in enumerate(lotes):
            future = executor.submit(procesar_lote_funcionarios, lote, fecha_corte, lote_id)
            future_to_lote[future] = lote_id
        
        # Recopilar resultados
        for future in as_completed(future_to_lote):
            lote_id = future_to_lote[future]
            try:
                resultado_lote = future.result()
                funcionarios_procesados.extend(resultado_lote)
                logger.info(f"[MULTIHILO] Lote {lote_id} completado - {len(resultado_lote)} funcionarios")
            except Exception as e:
                logger.error(f"[MULTIHILO] Error en lote {lote_id}: {e}")
    
    total_time = time.time() - start_time
    velocidad = len(funcionarios_procesados) / total_time if total_time > 0 else 0
    
    logger.info(f"[MULTIHILO] Procesamiento completado en {total_time:.2f}s")
    logger.info(f"[MULTIHILO] Velocidad: {velocidad:.2f} funcionarios/segundo")
    logger.info(f"[MULTIHILO] Total procesados: {len(funcionarios_procesados)}/{total_funcionarios}")
    
    return funcionarios_procesados

@router.post("/ascenso/generar-listado")
async def generar_listado_ascenso(request: ListadoAscensoRequest) -> ListadoAscensoResponse:
    """
    Genera el listado de ascensos según la fecha de corte especificada.
    Utiliza procesamiento multihilo y por lotes para optimizar el rendimiento.
    """
    try:
        logger.info(f"Generando listado de ascensos para fecha: {request.fecha_corte}")
        logger.info(f"Configuración: MAX_WORKERS={MAX_WORKERS}, BATCH_SIZE={BATCH_SIZE}")
        
        # Obtener funcionarios de la base de datos
        funcionarios_data = obtener_funcionarios_para_ascenso(request.fecha_corte)
        
        logger.info(f"Obtenidos {len(funcionarios_data)} funcionarios para evaluar")
        
        # Procesar funcionarios usando multihilo y lotes
        if len(funcionarios_data) >= BATCH_SIZE:
            logger.info("Usando procesamiento multihilo optimizado")
            funcionarios_procesados = procesar_funcionarios_multihilo(funcionarios_data, request.fecha_corte)
        else:
            logger.info("Usando procesamiento secuencial (pocos funcionarios)")
            funcionarios_procesados = []
            for funcionario_data in funcionarios_data:
                try:
                    funcionario_procesado = procesar_funcionario_para_ascenso(funcionario_data, request.fecha_corte)
                    funcionarios_procesados.append(funcionario_procesado)
                except Exception as e:
                    logger.error(f"Error procesando funcionario {funcionario_data.get('cedula', 'N/A')}: {e}")
                    continue
        
        logger.info(f"Procesados {len(funcionarios_procesados)} funcionarios exitosamente")
        
        # Organizar en listas
        listas = organizar_listas_ascenso(funcionarios_procesados)
        
        # Calcular estadísticas
        estadisticas = {
            "total_evaluados": len(funcionarios_procesados),
            "cumple_todos_requisitos": len(listas["cumple_todos_requisitos"]),
            "falta_nivel_academico": len(listas["falta_nivel_academico"]),
            "falta_tiempo_rango": len(listas["falta_tiempo_rango"]),
            "falta_tiempo_servicio": len(listas["falta_tiempo_servicio"]),
            "expediente_cerrado_reciente": len(listas["expediente_cerrado_reciente"]),
            "expediente_abierto": len(listas["expediente_abierto"]),
            "condicion_actual_invalida": len(listas["condicion_actual_invalida"])
        }
        
        # Generar archivo Excel
        try:
            archivo_excel = generar_excel_ascensos(listas, request.fecha_corte)
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            archivo_excel = None
        
        logger.info(f"Listado generado exitosamente. Estadísticas: {estadisticas}")
        
        return ListadoAscensoResponse(
            fecha_corte=request.fecha_corte,
            total_funcionarios_evaluados=len(funcionarios_procesados),
            listas=listas,
            estadisticas=estadisticas,
            archivo_excel=archivo_excel
        )
        
    except Exception as e:
        logger.error(f"Error generando listado de ascensos: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando listado: {str(e)}")

@router.get("/ascenso/funcionarios-por-rango")
async def get_funcionarios_por_rango(rango: str = Query(..., description="Rango a consultar")):
    """
    Obtiene la cantidad de funcionarios por rango específico.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = """
        SELECT 
            COUNT(*) as total,
            rango_actual
        FROM funcionarios 
        WHERE 
            status = 'ACTIVO'
            AND UPPER(tipo) = 'UNIFORMADO'
            AND UPPER(COALESCE(condicion_actual, '')) NOT IN ('SOLICITADO', 'DESTITUIDO', 'PRIVADO DE LIBERTAD')
            AND UPPER(rango_actual) = UPPER(%s)
        GROUP BY rango_actual
        """
        
        cursor.execute(query, (rango,))
        resultado = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if resultado:
            return {
                "rango": resultado['rango_actual'],
                "total_funcionarios": resultado['total'],
                "criterios_ascenso": CRITERIOS_ASCENSO.get(rango.upper(), {})
            }
        else:
            return {
                "rango": rango,
                "total_funcionarios": 0,
                "criterios_ascenso": CRITERIOS_ASCENSO.get(rango.upper(), {})
            }
            
    except Exception as e:
        logger.error(f"Error obteniendo funcionarios por rango: {e}")
        raise HTTPException(status_code=500, detail=f"Error consultando funcionarios: {str(e)}")

@router.get("/ascenso/estadisticas-generales")
async def get_estadisticas_generales():
    """
    Obtiene estadísticas generales de funcionarios por rango.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        query = """
        SELECT 
            rango_actual,
            COUNT(*) as total,
            COUNT(CASE WHEN sexo = 'M' THEN 1 END) as masculino,
            COUNT(CASE WHEN sexo = 'F' THEN 1 END) as femenino,
            AVG(EXTRACT(YEAR FROM AGE(CURRENT_DATE, fecha_nacimiento))) as edad_promedio,
            AVG(EXTRACT(YEAR FROM AGE(CURRENT_DATE, fecha_ingreso))) as antiguedad_promedio
        FROM funcionarios 
        WHERE 
            status = 'ACTIVO'
            AND UPPER(tipo) = 'UNIFORMADO'
            AND UPPER(COALESCE(condicion_actual, '')) NOT IN ('SOLICITADO', 'DESTITUIDO', 'PRIVADO DE LIBERTAD')
            AND rango_actual IS NOT NULL
            AND rango_actual != ''
        GROUP BY rango_actual
        ORDER BY 
            CASE rango_actual
                WHEN 'AGENTE' THEN 1
                WHEN 'OFICIAL' THEN 2
                WHEN 'PRIMER OFICIAL' THEN 3
                WHEN 'OFICIAL JEFE' THEN 4
                WHEN 'INSPECTOR' THEN 5
                WHEN 'PRIMER INSPECTOR' THEN 6
                WHEN 'INSPECTOR JEFE' THEN 7
                WHEN 'COMISARIO' THEN 8
                WHEN 'PRIMER COMISARIO' THEN 9
                WHEN 'COMISARIO JEFE' THEN 10
                ELSE 11
            END
        """
        
        cursor.execute(query)
        resultados = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        estadisticas = []
        total_general = 0
        
        for resultado in resultados:
            total_general += resultado['total']
            estadisticas.append({
                "rango": resultado['rango_actual'],
                "total": resultado['total'],
                "masculino": resultado['masculino'] or 0,
                "femenino": resultado['femenino'] or 0,
                "edad_promedio": round(float(resultado['edad_promedio'] or 0), 1),
                "antiguedad_promedio": round(float(resultado['antiguedad_promedio'] or 0), 1),
                "criterios_ascenso": CRITERIOS_ASCENSO.get(resultado['rango_actual'], {})
            })
        
        return {
            "total_funcionarios": total_general,
            "estadisticas_por_rango": estadisticas,
            "fecha_consulta": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas generales: {e}")
        raise HTTPException(status_code=500, detail=f"Error obteniendo estadísticas: {str(e)}")

@router.post("/ascenso/simular-ascenso")
async def simular_ascenso(cedula: str, fecha_simulacion: date):
    """
    Simula el ascenso de un funcionario específico en una fecha determinada.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Obtener datos del funcionario específico
        query = """
        SELECT 
            f.id,
            f.cedula,
            f.nombre_completo,
            f.sexo,
            f.fecha_nacimiento,
            f.fecha_ingreso,
            f.fecha_ultimo_ascenso,
            f.rango_actual,
            f.status,
            f.condicion_actual,
            f.tipo,
            f.grado_instruccion,
            
            -- Antecedentes académicos
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'grado_instruccion', aa.grado_instruccion,
                        'institucion', aa.institucion,
                        'fecha_graduacion', aa.fecha_graduacion
                    )
                ) FILTER (WHERE aa.id IS NOT NULL), 
                '[]'::json
            ) as antecedentes_academicos,
            
            -- Tiempo de servicio adicional
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'institucion', ts.institucion,
                        'fecha_ingreso', ts.fecha_ingreso,
                        'fecha_egreso', ts.fecha_egreso,
                        'cargo', ts.cargo
                    )
                ) FILTER (WHERE ts.id IS NOT NULL), 
                '[]'::json
            ) as tiempos_servicio,
            
            -- Expedientes
            COALESCE(
                json_agg(
                    DISTINCT jsonb_build_object(
                        'nro_expediente', e.nro_expediente,
                        'fecha_inicio', e.fecha_inicio,
                        'fecha_finalizacion', e.fecha_finalizacion,
                        'estatus', e.estatus,
                        'falta', e.falta,
                        'decision', e.decision
                    )
                ) FILTER (WHERE e.id_exp IS NOT NULL), 
                '[]'::json
            ) as expedientes
            
        FROM funcionarios f
        LEFT JOIN antecedentes_academicos aa ON f.id = aa.funcionario_id
        LEFT JOIN tiempo_servicio ts ON f.id = ts.funcionario_id
        LEFT JOIN expedientes e ON f.id = e.funcionario_id
        WHERE f.cedula = %s
        GROUP BY f.id
        """
        
        cursor.execute(query, (cedula,))
        funcionario_data = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if not funcionario_data:
            raise HTTPException(status_code=404, detail="Funcionario no encontrado")
        
        # Procesar funcionario para la fecha de simulación
        funcionario_procesado = procesar_funcionario_para_ascenso(dict(funcionario_data), fecha_simulacion)
        
        return {
            "funcionario": funcionario_procesado,
            "fecha_simulacion": fecha_simulacion,
            "es_elegible": funcionario_procesado.cumple_todos_requisitos,
            "recomendaciones": {
                "puede_ascender": funcionario_procesado.cumple_todos_requisitos,
                "requisitos_faltantes": funcionario_procesado.observaciones if not funcionario_procesado.cumple_todos_requisitos else "Ninguno",
                "puntos_merito": funcionario_procesado.total_puntos,
                "rango_destino": funcionario_procesado.rango_a_aplicar
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error simulando ascenso: {e}")
        raise HTTPException(status_code=500, detail=f"Error en simulación: {str(e)}")

@router.get("/ascenso/exportar-excel/{fecha_corte}")
async def exportar_excel_ascensos(fecha_corte: date):
    """
    Exporta directamente el archivo Excel de ascensos para una fecha específica.
    """
    try:
        # Crear request para generar listado
        request = ListadoAscensoRequest(fecha_corte=fecha_corte)
        
        # Generar listado
        listado = await generar_listado_ascenso(request)
        
        if not listado.archivo_excel:
            raise HTTPException(status_code=500, detail="No se pudo generar el archivo Excel")
        
        # Decodificar el archivo base64
        excel_content = base64.b64decode(listado.archivo_excel)
        
        # Crear respuesta con el archivo
        from fastapi.responses import Response
        
        filename = f"listado_ascensos_{fecha_corte.strftime('%Y_%m_%d')}.xlsx"
        
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exportando Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error exportando archivo: {str(e)}")

@router.post("/ascenso/cargar-excel-evaluado")
async def cargar_excel_evaluado(file: UploadFile = File(...)):
    """
    Carga un archivo Excel con las evaluaciones de ascenso completadas y actualiza el historial de ascensos.
    """
    try:
        from fastapi import UploadFile, File
        import tempfile
        import os
        
        logger.info(f"Cargando archivo Excel evaluado: {file.filename}")
        
        # Validar que sea un archivo Excel
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls)")
        
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name
        
        try:
            # Leer el Excel
            df = pd.read_excel(temp_path)
            
            # Validar columnas requeridas
            required_columns = ['Cédula', 'Apto', 'No Apto']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Faltan las siguientes columnas en el Excel: {', '.join(missing_columns)}"
                )
            
            # Procesar las evaluaciones
            evaluaciones_procesadas = []
            evaluaciones_con_error = []
            
            conn = get_db_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            for index, row in df.iterrows():
                try:
                    cedula = str(row['Cédula']).strip()
                    apto = str(row['Apto']).strip().upper() if pd.notna(row['Apto']) else ''
                    no_apto = str(row['No Apto']).strip().upper() if pd.notna(row['No Apto']) else ''
                    
                    # Validar que solo una opción esté marcada
                    if apto == 'SÍ' and no_apto == 'SÍ':
                        evaluaciones_con_error.append({
                            'fila': index + 2,
                            'cedula': cedula,
                            'error': 'No puede estar marcado como Apto y No Apto al mismo tiempo'
                        })
                        continue
                    
                    if apto != 'SÍ' and no_apto != 'SÍ':
                        # Saltar filas sin evaluación
                        continue
                    
                    # Determinar el estado
                    estado_evaluacion = 'APTO' if apto == 'SÍ' else 'NO_APTO'
                    
                    # Buscar funcionario por cédula
                    cursor.execute("SELECT id FROM funcionarios WHERE cedula = %s", (cedula,))
                    funcionario = cursor.fetchone()
                    
                    if not funcionario:
                        evaluaciones_con_error.append({
                            'fila': index + 2,
                            'cedula': cedula,
                            'error': 'Funcionario no encontrado'
                        })
                        continue
                    
                    # Insertar o actualizar en historial de ascensos
                    insert_query = """
                    INSERT INTO historial_ascensos 
                    (funcionario_id, rango_anterior, rango_nuevo, tipo_ascenso, fecha_ascenso, decision, estado, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    # Obtener rango actual del funcionario
                    cursor.execute("SELECT rango_actual FROM funcionarios WHERE id = %s", (funcionario['id'],))
                    funcionario_data = cursor.fetchone()
                    rango_actual = funcionario_data['rango_actual'] if funcionario_data else 'N/A'
                    
                    # Determinar rango siguiente
                    criterios = CRITERIOS_ASCENSO.get(rango_actual.upper(), {})
                    rango_siguiente = criterios.get('siguienteRango', 'N/A')
                    
                    cursor.execute(insert_query, (
                        funcionario['id'],
                        rango_actual,
                        rango_siguiente if estado_evaluacion == 'APTO' else rango_actual,
                        'EVALUACION_ASCENSO',
                        datetime.now().date(),
                        f'Evaluación de ascenso: {estado_evaluacion}',
                        'APROBADO' if estado_evaluacion == 'APTO' else 'RECHAZADO',
                        f'Evaluación cargada desde Excel el {datetime.now().strftime("%d-%m-%Y")}'
                    ))
                    
                    evaluaciones_procesadas.append({
                        'cedula': cedula,
                        'estado': estado_evaluacion,
                        'rango_actual': rango_actual,
                        'rango_siguiente': rango_siguiente if estado_evaluacion == 'APTO' else rango_actual
                    })
                    
                except Exception as row_error:
                    evaluaciones_con_error.append({
                        'fila': index + 2,
                        'cedula': cedula if 'cedula' in locals() else 'N/A',
                        'error': str(row_error)
                    })
                    continue
            
            # Confirmar transacción
            conn.commit()
            cursor.close()
            conn.close()
            
            logger.info(f"Procesadas {len(evaluaciones_procesadas)} evaluaciones exitosamente")
            if evaluaciones_con_error:
                logger.warning(f"Se encontraron {len(evaluaciones_con_error)} errores durante el procesamiento")
            
            return {
                "success": True,
                "message": f"Archivo procesado exitosamente. {len(evaluaciones_procesadas)} evaluaciones cargadas.",
                "evaluaciones_procesadas": len(evaluaciones_procesadas),
                "evaluaciones_con_error": len(evaluaciones_con_error),
                "errores": evaluaciones_con_error[:10] if evaluaciones_con_error else [],  # Mostrar solo los primeros 10 errores
                "detalles_procesadas": evaluaciones_procesadas[:5]  # Mostrar solo las primeras 5 para referencia
            }
            
        finally:
            # Limpiar archivo temporal
            if os.path.exists(temp_path):
                os.unlink(temp_path)
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error cargando Excel evaluado: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")
